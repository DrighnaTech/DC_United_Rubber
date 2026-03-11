"""Excel Builder — creates multi-sheet workbook matching the analysis table format."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styles
_thin = Side(style="thin", color="B0BEC5")
BD = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
FT = Font(name="Arial", bold=True, size=14, color="FFFFFF")
FH = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FL = Font(name="Arial", bold=True, size=9, color="1F4E79")
FV = Font(name="Arial", size=9, color="000000")
FN = Font(name="Arial", italic=True, size=8, color="888888")
FG = Font(name="Arial", bold=True, size=10, color="1B5E20")
FR = Font(name="Arial", bold=True, size=10, color="C62828")

COLORS = {
    "title_block":       "1F4E79",
    "dimensions":        "006064",
    "coordinate_points": "2E75B6",
    "bom":               "BF6900",
    "notes":             "5E35B1",
    "standards":         "1B5E20",
    "general_tolerances":"0D9488",
    "marking_table":     "7B2D8E",
    "revision_history":  "795548",
    "derived_data":      "C62828",
    "costing_input":     "0D47A1",
    "accuracy":          "37474F",
}

def _c(ws, r, c, val="", font=FV, fill=None, align=AL):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font; cell.alignment = align; cell.border = BD
    if fill: cell.fill = fill
    return cell

def _header_row(ws, row, headers, color):
    fill = PatternFill("solid", fgColor=color)
    for i, h in enumerate(headers, 1):
        _c(ws, row, i, h, font=FH, fill=fill, align=AC)

def _title(ws, row, text, color, ncols):
    fill = PatternFill("solid", fgColor=color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    _c(ws, row, 1, text, font=FT, fill=fill, align=AL)
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = BD
    ws.row_dimensions[row].height = 30

def _conf_font(conf):
    try:
        v = int(conf) if conf else 0
    except (ValueError, TypeError):
        v = 0
    if v >= 95: return Font(name="Arial", bold=True, size=9, color="1B5E20")
    if v >= 80: return Font(name="Arial", size=9, color="E65100")
    return Font(name="Arial", size=9, color="C62828")

def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _safe(data, key, default=""):
    v = data.get(key, default)
    return str(v) if v is not None else ""


def build_excel(data, pdf_name):
    wb = Workbook()
    info = data.get("drawing_info", {})
    dwg = info.get("drawing_number", pdf_name)

    # ═══════════════════════════════════════════════════════════
    # SHEET 1: Title Block
    # ═══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Title Block"
    ws.sheet_properties.tabColor = COLORS["title_block"]
    _set_widths(ws, [5, 8, 30, 50, 12, 30])
    _title(ws, 1, f"  TITLE BLOCK — {dwg}", COLORS["title_block"], 6)
    headers = ["#", "Num", "Field", "Extracted Value", "Confidence", "Source Location"]
    _header_row(ws, 2, headers, COLORS["title_block"])
    rows = data.get("title_block", [])
    for i, r in enumerate(rows, 1):
        row = i + 2
        _c(ws, row, 1, i, font=FN, align=AC)
        _c(ws, row, 2, r.get("num", i), font=FN, align=AC)
        _c(ws, row, 3, _safe(r, "field"), font=FL)
        _c(ws, row, 4, _safe(r, "value"), font=FV)
        conf = r.get("confidence", "")
        _c(ws, row, 5, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
        _c(ws, row, 6, _safe(r, "source_location"), font=FN)
    ws.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 2: Dimensions
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Dimensions")
    ws2.sheet_properties.tabColor = COLORS["dimensions"]
    _set_widths(ws2, [8, 14, 34, 12, 10, 10, 8, 24, 18, 12])
    _title(ws2, 1, f"  DIMENSIONS — {dwg}", COLORS["dimensions"], 10)
    hdrs = ["Dim ID", "Category", "Feature", "Nominal", "Tol+", "Tol-", "Unit", "Raw Text", "View", "Confidence"]
    _header_row(ws2, 2, hdrs, COLORS["dimensions"])
    dims = data.get("dimensions", [])
    for i, d in enumerate(dims):
        row = i + 3
        _c(ws2, row, 1, _safe(d, "dim_id"), font=FL, align=AC)
        _c(ws2, row, 2, _safe(d, "category"), font=FN)
        _c(ws2, row, 3, _safe(d, "feature"), font=FV)
        _c(ws2, row, 4, _safe(d, "nominal"), font=Font(name="Arial", bold=True, size=10), align=AC)
        _c(ws2, row, 5, _safe(d, "tol_plus") or "—", font=FV, align=AC)
        _c(ws2, row, 6, _safe(d, "tol_minus") or "—", font=FV, align=AC)
        _c(ws2, row, 7, _safe(d, "unit"), font=FN, align=AC)
        _c(ws2, row, 8, _safe(d, "raw_text"), font=FN)
        _c(ws2, row, 9, _safe(d, "view"), font=FN)
        conf = d.get("confidence", "")
        _c(ws2, row, 10, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    ws2.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 3: Coordinate Points
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Coordinates")
    ws3.sheet_properties.tabColor = COLORS["coordinate_points"]
    _set_widths(ws3, [10, 14, 14, 14, 8, 12])
    _title(ws3, 1, f"  COORDINATE POINTS — {dwg}", COLORS["coordinate_points"], 6)
    coords = data.get("coordinate_points", [])
    if coords:
        _header_row(ws3, 2, ["Point", "X", "Y", "Z", "Unit", "Confidence"], COLORS["coordinate_points"])
        for i, cp in enumerate(coords):
            row = i + 3
            _c(ws3, row, 1, _safe(cp, "point"), font=FL, align=AC)
            _c(ws3, row, 2, cp.get("x", ""), font=Font(name="Arial", bold=True, size=10), align=AC)
            _c(ws3, row, 3, cp.get("y", ""), font=Font(name="Arial", bold=True, size=10), align=AC)
            _c(ws3, row, 4, cp.get("z", ""), font=Font(name="Arial", bold=True, size=10), align=AC)
            _c(ws3, row, 5, _safe(cp, "unit"), font=FN, align=AC)
            conf = cp.get("confidence", "")
            _c(ws3, row, 6, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    else:
        _c(ws3, 2, 1, "No coordinate data — expected for flat/symmetric parts", font=FN)
    ws3.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 4: BOM
    # ═══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("BOM")
    ws4.sheet_properties.tabColor = COLORS["bom"]
    _set_widths(ws4, [8, 20, 8, 36, 20, 14, 12])
    _title(ws4, 1, f"  BILL OF MATERIALS — {dwg}", COLORS["bom"], 7)
    bom = data.get("bom", [])
    if bom:
        _header_row(ws4, 2, ["Item", "Part Number", "Rev", "Description", "Matl Spec", "Matl Qty", "Confidence"], COLORS["bom"])
        for i, b in enumerate(bom):
            row = i + 3
            _c(ws4, row, 1, b.get("item", ""), font=FV, align=AC)
            _c(ws4, row, 2, _safe(b, "part_number"), font=FL)
            _c(ws4, row, 3, _safe(b, "rev"), font=FV, align=AC)
            _c(ws4, row, 4, _safe(b, "description"), font=FV)
            _c(ws4, row, 5, _safe(b, "matl_spec"), font=FV)
            _c(ws4, row, 6, _safe(b, "matl_qty"), font=FV, align=AC)
            conf = b.get("confidence", "")
            _c(ws4, row, 7, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    else:
        _c(ws4, 2, 1, "No BOM — single-piece component", font=FN)
    ws4.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 5: Notes
    # ═══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Notes")
    ws5.sheet_properties.tabColor = COLORS["notes"]
    _set_widths(ws5, [10, 16, 60, 60, 12])
    _title(ws5, 1, f"  NOTES & CALLOUTS — {dwg}", COLORS["notes"], 5)
    notes = data.get("notes", [])
    _header_row(ws5, 2, ["Note #", "Category", "Full Text", "English Translation", "Confidence"], COLORS["notes"])
    for i, n in enumerate(notes):
        row = i + 3
        _c(ws5, row, 1, _safe(n, "note_num"), font=FN, align=AC)
        _c(ws5, row, 2, _safe(n, "category"), font=FL)
        _c(ws5, row, 3, _safe(n, "full_text"), font=FV)
        _c(ws5, row, 4, _safe(n, "english_translation"), font=FN)
        conf = n.get("confidence", "")
        _c(ws5, row, 5, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    ws5.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 6: Standards
    # ═══════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Standards")
    ws6.sheet_properties.tabColor = COLORS["standards"]
    _set_widths(ws6, [6, 16, 24, 50, 12])
    _title(ws6, 1, f"  STANDARDS & REFERENCES — {dwg}", COLORS["standards"], 5)
    stds = data.get("standards", [])
    _header_row(ws6, 2, ["#", "Type", "Code", "Context / Description", "Confidence"], COLORS["standards"])
    for i, s in enumerate(stds):
        row = i + 3
        _c(ws6, row, 1, s.get("num", i + 1), font=FN, align=AC)
        _c(ws6, row, 2, _safe(s, "standard_type"), font=FL)
        _c(ws6, row, 3, _safe(s, "code"), font=Font(name="Arial", bold=True, size=9))
        _c(ws6, row, 4, _safe(s, "context"), font=FV)
        conf = s.get("confidence", "")
        _c(ws6, row, 5, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    ws6.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 6B: General Tolerances (Undimensioned)
    # ═══════════════════════════════════════════════════════════
    gt_rows = data.get("general_tolerances", [])
    gt_note = data.get("general_tolerance_note", "")
    if gt_rows or gt_note:
        wsgt = wb.create_sheet("General Tolerances")
        wsgt.sheet_properties.tabColor = COLORS["general_tolerances"]
        _set_widths(wsgt, [22, 14, 14, 10, 18, 12])
        _title(wsgt, 1, f"  GENERAL TOLERANCES (UNDIMENSIONED) — {dwg}", COLORS["general_tolerances"], 6)
        _header_row(wsgt, 2, ["Size Range", "Tol +", "Tol −", "Unit", "Tol Class", "Confidence"], COLORS["general_tolerances"])
        gt_row = 3
        for gt in gt_rows:
            _c(wsgt, gt_row, 1, _safe(gt, "range"), font=Font(name="Arial", bold=True, size=9))
            _c(wsgt, gt_row, 2, _safe(gt, "tol_plus"), font=Font(name="Arial", bold=True, size=9, color="1B5E20"), align=AC)
            _c(wsgt, gt_row, 3, _safe(gt, "tol_minus"), font=Font(name="Arial", bold=True, size=9, color="C62828"), align=AC)
            _c(wsgt, gt_row, 4, _safe(gt, "unit") or "mm", font=FV, align=AC)
            _c(wsgt, gt_row, 5, _safe(gt, "tolerance_class"), font=FV, align=AC)
            conf = gt.get("confidence", "")
            _c(wsgt, gt_row, 6, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
            gt_row += 1
        if gt_note:
            gt_row += 1
            wsgt.merge_cells(start_row=gt_row, start_column=1, end_row=gt_row, end_column=6)
            _c(wsgt, gt_row, 1, "TOLERANCE NOTE:", font=Font(name="Arial", bold=True, size=9, color="0D47A1"))
            gt_row += 1
            wsgt.merge_cells(start_row=gt_row, start_column=1, end_row=gt_row + 2, end_column=6)
            cell = wsgt.cell(row=gt_row, column=1, value=str(gt_note))
            cell.font = Font(name="Arial", italic=True, size=9, color="333333")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            wsgt.row_dimensions[gt_row].height = 60
        wsgt.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 7: Marking Table
    # ═══════════════════════════════════════════════════════════
    marks = data.get("marking_table", [])
    if marks:
        ws7 = wb.create_sheet("Marking Table")
        ws7.sheet_properties.tabColor = COLORS["marking_table"]
        _set_widths(ws7, [28, 28, 18, 40, 12])
        _title(ws7, 1, f"  KENNZEICHNUNG / MARKING — {dwg}", COLORS["marking_table"], 5)
        _header_row(ws7, 2, ["German (Original)", "English Translation", "Standard", "Description", "Confidence"], COLORS["marking_table"])
        for i, m in enumerate(marks):
            row = i + 3
            german = m.get("marking_type_german", m.get("marking_type", ""))
            english = m.get("marking_type", german)
            _c(ws7, row, 1, german, font=Font(name="Arial", bold=True, size=9, color="7B2D8E"))
            _c(ws7, row, 2, english, font=FL)
            _c(ws7, row, 3, _safe(m, "standard"), font=Font(name="Arial", bold=True, size=9))
            _c(ws7, row, 4, _safe(m, "description"), font=FV)
            conf = m.get("confidence", "")
            _c(ws7, row, 5, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
        ws7.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 8: Revision History
    # ═══════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Revision History")
    ws8.sheet_properties.tabColor = COLORS["revision_history"]
    _set_widths(ws8, [8, 10, 36, 36, 14, 18, 18, 12])
    _title(ws8, 1, f"  REVISION HISTORY — {dwg}", COLORS["revision_history"], 8)
    revs = data.get("revision_history", [])
    _header_row(ws8, 2, ["Rev", "Section", "Description", "English", "Date", "Change Code", "Approved By", "Confidence"], COLORS["revision_history"])
    for i, r in enumerate(revs):
        row = i + 3
        _c(ws8, row, 1, _safe(r, "rev"), font=FL, align=AC)
        _c(ws8, row, 2, _safe(r, "section"), font=FN, align=AC)
        _c(ws8, row, 3, _safe(r, "description"), font=FV)
        _c(ws8, row, 4, _safe(r, "english_translation"), font=FN)
        _c(ws8, row, 5, _safe(r, "date"), font=FV, align=AC)
        _c(ws8, row, 6, _safe(r, "change_code"), font=FN, align=AC)
        _c(ws8, row, 7, _safe(r, "approved_by"), font=FV)
        conf = r.get("confidence", "")
        _c(ws8, row, 8, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    ws8.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 9: Derived Engineering Data
    # ═══════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("Derived Data")
    ws9.sheet_properties.tabColor = COLORS["derived_data"]
    _set_widths(ws9, [30, 30, 46, 12])
    _title(ws9, 1, f"  DERIVED ENGINEERING DATA — {dwg}", COLORS["derived_data"], 4)
    dd = data.get("derived_data", [])
    _header_row(ws9, 2, ["Parameter", "Value", "How Derived", "Confidence"], COLORS["derived_data"])
    for i, d in enumerate(dd):
        row = i + 3
        _c(ws9, row, 1, _safe(d, "parameter"), font=FL)
        _c(ws9, row, 2, _safe(d, "value"), font=Font(name="Arial", bold=True, size=10))
        _c(ws9, row, 3, _safe(d, "how_derived"), font=FN)
        conf = d.get("confidence", "")
        _c(ws9, row, 4, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    ws9.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 10: Costing Input
    # ═══════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("Costing Input")
    ws10.sheet_properties.tabColor = COLORS["costing_input"]
    _set_widths(ws10, [30, 36, 30, 12])
    _title(ws10, 1, f"  COSTING INPUT DATA — {dwg}", COLORS["costing_input"], 4)
    ci = data.get("costing_input", [])
    _header_row(ws10, 2, ["Parameter", "Value", "Source", "Confidence"], COLORS["costing_input"])
    for i, c in enumerate(ci):
        row = i + 3
        _c(ws10, row, 1, _safe(c, "parameter"), font=FL)
        _c(ws10, row, 2, _safe(c, "value"), font=Font(name="Arial", bold=True, size=10))
        _c(ws10, row, 3, _safe(c, "source"), font=FN)
        conf = c.get("confidence", "")
        _c(ws10, row, 4, f"{conf}%" if conf else "", font=_conf_font(conf), align=AC)
    ws10.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # SHEET 11: Accuracy Summary
    # ═══════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("Accuracy")
    ws11.sheet_properties.tabColor = COLORS["accuracy"]
    _set_widths(ws11, [34, 20, 14])
    _title(ws11, 1, f"  EXTRACTION ACCURACY — {dwg}", COLORS["accuracy"], 3)

    acc = data.get("accuracy_summary", {})
    summary_rows = [
        ("Total Data Points Extracted", acc.get("total_data_points", "—")),
        ("100% Confidence (directly readable)", acc.get("confidence_100", acc.get("confidence_100_pct", "—"))),
        ("90-99% Confidence (clear + context)", acc.get("confidence_90_99", acc.get("confidence_90_99_pct", "—"))),
        ("80-89% Confidence (inferred)", acc.get("confidence_80_89", acc.get("confidence_80_89_pct", "—"))),
        ("<80% Confidence (assumption)", acc.get("confidence_below_80", acc.get("confidence_below_80_pct", "—"))),
        ("Overall Extraction Accuracy", f"{acc.get('overall_accuracy_pct', '—')}%"),
    ]
    _header_row(ws11, 2, ["Metric", "Value", ""], COLORS["accuracy"])
    for i, (label, val) in enumerate(summary_rows):
        row = i + 3
        _c(ws11, row, 1, label, font=FL)
        is_overall = "Overall" in label
        _c(ws11, row, 2, val,
           font=Font(name="Arial", bold=True, size=13 if is_overall else 11,
                     color="1B5E20" if is_overall else "000000"),
           align=AC)

    # Section breakdown
    row = len(summary_rows) + 4
    _c(ws11, row, 1, "SECTION BREAKDOWN", font=FH,
       fill=PatternFill("solid", fgColor="ECEFF1"), align=AL)
    _c(ws11, row, 2, "Count", font=FH,
       fill=PatternFill("solid", fgColor="ECEFF1"), align=AC)
    row += 1
    sections = [
        ("Title Block", len(data.get("title_block", []))),
        ("Dimensions", len(data.get("dimensions", []))),
        ("Coordinate Points", len(data.get("coordinate_points", []))),
        ("BOM Items", len(data.get("bom", []))),
        ("Notes & Callouts", len(data.get("notes", []))),
        ("Standards & References", len(data.get("standards", []))),
        ("General Tolerances", len(data.get("general_tolerances", []))),
        ("Marking Table", len(data.get("marking_table", []))),
        ("Revision History", len(data.get("revision_history", []))),
        ("Derived Engineering Data", len(data.get("derived_data", []))),
        ("Costing Input", len(data.get("costing_input", []))),
    ]
    total = 0
    for label, cnt in sections:
        _c(ws11, row, 1, label, font=FV)
        _c(ws11, row, 2, cnt, font=Font(name="Arial", bold=True, size=10), align=AC)
        total += cnt
        row += 1
    _c(ws11, row, 1, "TOTAL", font=Font(name="Arial", bold=True, size=11))
    _c(ws11, row, 2, total, font=Font(name="Arial", bold=True, size=12, color="1F4E79"), align=AC)

    ws11.freeze_panes = "A3"

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
